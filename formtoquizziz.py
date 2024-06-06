import docx
import re
import openpyxl

class Question: 
    def __init__(self, question):
        self.question = question
        self.answer = []
        self.rightAns = -1

    def addAnswer(self, answer):
        self.answer.append(answer)

    def setRightAns(self, rightAns):
        self.rightAns = rightAns

    def setRightAnsStr(self, string):
        self.rightAns = self.answer.index(string) + 1

    def getRightAns(self):
        return self.rightAns

    def getAns(self):
        return self.answer

    def getAnsLen(self):
        return len(self.answer)
    
    def getQuest(self):
        return self.question
    
    def __str__(self):
        return "Question: " + self.question + "\nAnswer: " + str(self.answer) + "\nRight Answer: " + str(self.rightAns) + "\n"

class Converter:
    def __init__(self, docfile, xlsxfile):
        self.question = []

        self.docfile = docfile
        self.xlsxfile = xlsxfile

        self.docf = docx.Document(self.docfile)
        self.workbook = openpyxl.load_workbook(self.xlsxfile)
        self.worksheet = self.workbook["Create a Quiz"]

    def readtxt(self):
        lineIndex = 0
        numOfLine = len(self.docf. paragraphs)

        paragraphs = []

        while(lineIndex < numOfLine):
            curLine = self.docf.paragraphs[lineIndex]
            lineText = re.sub(r'\s+', ' ', curLine.text.strip())

            if (len(curLine.text) <= 1 or curLine.text == "25/25" or curLine.text == "0/25" or curLine.text == "Câu trả lời đúng"):
                lineIndex += 1
                continue

            paragraphs.append(curLine)    

            lineIndex += 1

        lineIndex = 0
        
        while (lineIndex < len(paragraphs)):
            lineText = paragraphs[lineIndex]
            question = Question(lineText.text)

            # print(lineText.runs[0].font.color.rgb, lineText.text)
            # for run in lineText.runs:
                # print(run.font.color.rgb)

            if (str(lineText.runs[0].font.color.rgb) == "1E8E3E"):
                for i in range(4):
                    lineIndex += 1
                    curLine = paragraphs[lineIndex]
                    lineText = re.sub(r'\s+', ' ', curLine.text.strip())
                    
                    question.addAnswer(lineText)
                    if (str(curLine.runs[0].font.color.rgb) == "202124"):
                        question.setRightAns(i + 1)

            elif (str(lineText.runs[0].font.color.rgb) == "D93025"):
                for i in range(4):
                    lineIndex += 1
                    curLine = paragraphs[lineIndex]
                    lineText = re.sub(r'\s+', ' ', curLine.text.strip())

                    question.addAnswer(lineText)

                lineIndex += 1
                curLine = paragraphs[lineIndex]
                lineText = re.sub(r'\s+', ' ', curLine.text.strip())
                question.setRightAnsStr(lineText)
                # tmp = 

            lineIndex += 1
            print(question)
            self.question.append(question)

    def addSheet(self):
        row_num = 3

        for quest in self.question:
            self.worksheet.cell(row=row_num, column=1).value = quest.getQuest()
            self.worksheet.cell(row=row_num, column=2).value = 'Multiple Choice'

            ansArr = quest.getAns()
            numOfAns = len(ansArr)
            for i in range(numOfAns):
                self.worksheet.cell(row=row_num, column=3 + i).value = ansArr[i]

            self.worksheet.cell(row=row_num, column=7).value = quest.getRightAns()
            self.worksheet.cell(row=row_num, column=8).value = 60
            row_num += 1

        self.workbook.save(self.xlsxfile) 
        print("done")

c = Converter("H:\\Download\\1.docx", "H:\\Download\\1.xlsx")
c.readtxt()
c.addSheet()